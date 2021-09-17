import openpyxl
from openpyxl.utils import datetime
from datetime import datetime


Dict = {}
book = openpyxl.load_workbook(r"C:\Users\Southville\Desktop\Book1.xlsx")
sheet = book.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)       # read value from excel sheet
# sheet.cell(row=2, column=2).value = "dania"   # write value on excel sheet
# sheet.cell(row=3, column=2).value = "fazeela"   # write value on excel sheet
# book.save(filename = r'C:\Users\Southville\Desktop\PythonDemo.xlsx')  # after writing something save the file
# print(sheet.cell(row=2, column=2).value)
# # ---print total no. of rows and columns in excel sheet
# print(sheet.max_row)
# print(sheet.max_column)
# # ---we can target cell with 2 ways:
# # ---sheet.cell(row=1, column=2).value
# # ---sheet['A4'].value
# print(sheet['A4'].value)
# # ---print all the values of excel sheet
arr = []
print(sheet.max_row+1)
for i in range(2, sheet.max_row+1):
    # if sheet.cell(row=i, column=1).value == "testcase3": # to print specific cell's values
    for j in range(2, sheet.max_column+1):  # start from column 2, we don't need to print 'testcase3' in output
        Dict[sheet.cell(row=1 , column=j).value] = sheet.cell(row=i, column=j).value
            # print(sheet.cell(row=i, column=j).value)
    arr.append(Dict.copy())
print(arr)

for i in arr:
    print(i["FirstName"])
    print(i["LastName"])
    print(i["Email"])
    print(i["FirstName"])
    print(i["MobileNumber"])
    print(i["Gender"])
    print(i["DateOfBirth"])
    date = i["DateOfBirth"]
    print(date.day)
    print(date.month)
    print(date.year)
    print("-----------------------")